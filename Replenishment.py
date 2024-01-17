from openpyxl import load_workbook  
import math
import tkinter as tk 
from tkinter import messagebox
from tkinter import filedialog

def calculate_totals(worksheet,totle_replenishment): 
    totle_sales =0
    totle_stock = 0
    totle_stock_on_road = 0
    totle_replenishment_final = 0
    # 计算D列的总和  
    for row in range(2, worksheet.max_row + 1):  
        cell = worksheet.cell(row=row, column=4)  # 读取D列的单元格  
        if cell.value is not None:  # 检查单元格是否为空  
            totle_sales += int(cell.value)  # 将值加到总和中

    # 计算C列的总和  
    for row in range(2, worksheet.max_row + 1):  
        cell = worksheet.cell(row=row, column=3)  # 读取c列的单元格  
        if cell.value is not None:  # 检查单元格是否为空  
            totle_stock_on_road += int(cell.value)  # 将值加到总和中

    # 计算B列的总和 
    for row in range(2, worksheet.max_row + 1):  
        cell = worksheet.cell(row=row, column=2)  # 读取b列的单元格  
        if cell.value is not None:  # 检查单元格是否为空  
            totle_stock += int(cell.value)  # 将值加到总和中 

    # 计算每一个产品的销售比例
    for row in range(2, worksheet.max_row + 1):  
        worksheet.cell(row=row, column=5, value=(worksheet.cell(row=row, column=4).value / totle_sales)) 
    
    for row in range(2, worksheet.max_row + 1):  
        replenishment = worksheet.cell(row=row, column=5).value * (totle_replenishment + totle_stock_on_road + totle_stock) - (worksheet.cell(row=row, column=2).value +worksheet.cell(row=row, column=3).value)
        worksheet.cell(row=row, column=6, value=math.ceil(replenishment))
    
    for row in range(2, worksheet.max_row + 1):  
        cell = worksheet.cell(row=row, column=6)  # 读取F列的单元格  
        if cell.value is not None:  # 检查单元格是否为空  
            totle_replenishment_final += int(cell.value)  # 将值加到总和中 
    worksheet.cell(row=worksheet.max_row+1, column=5, value="此次补货总量：")
    worksheet.cell(row=worksheet.max_row, column=6, value=totle_replenishment_final)


def calculate_totals_button():  
    #try:
        # 执行计算函数  
    totle_replenishment = int(replenishment.get())
    calculate_totals(worksheet,totle_replenishment)
    messagebox.showinfo('提示', '补货数量计算完成')
    # 保存更改并关闭工作簿  
    workbook.save(filename=file_path)
    #except Exception as e:  
        #messagebox.showerror(Exception, e)


def open_file_dialog():  
    global file_path,workbook, worksheet
    # 打开文件选择对话框  
    file_path = filedialog.askopenfilename()  
    print("选择的文件路径是:", file_path)
    # 打开excel文件  
    workbook = load_workbook(filename= file_path)  
    # 选择要操作的工作表，例如第一个工作表  
    worksheet = workbook['Sheet1'] 
    worksheet.cell(row=1, column=5, value='销售比例')
    worksheet.cell(row=1, column=6, value='补货数量')
    lable_path.config(text="当前选择文件为："+file_path)



root = tk.Tk()
#root.geometry('350x80') 
root.title("补货数量")
frame1 = tk.Frame(root)
frame1.pack(fill=tk.X)
lable = tk.Label(frame1, text='此次补货总数量:')
replenishment = tk.Entry(frame1)
lable.pack(side=tk.LEFT)
replenishment.pack(side=tk.LEFT)


frame = tk.Frame(root)
frame.pack(fill=tk.X)
open_button = tk.Button(frame, text="选择文件", command=open_file_dialog)
calculate_button = tk.Button(frame, text='计算补货数量', command=calculate_totals_button)
open_button.pack(side=tk.LEFT) 
calculate_button.pack(side=tk.LEFT)

frame3 = tk.Frame(root)
frame3.pack(fill=tk.X)
lable_path = tk.Label(frame3, text="当前选择文件为：")
lable_path.pack(side= tk.LEFT)

root.mainloop()

 
